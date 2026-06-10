#!/usr/bin/env python3
"""
Import nośników Optokom z arkusza "LISTA LOKALIZACJI 2026 REZERWACJE".

Parsuje XLSX, mapuje wiersze na pola modelu Advertisement, wylicza ceny
(reguła powierzchnia × technologia z cennika agencji), statusy (z kalendarza
rezerwacji), geokoduje braki GPS i zapisuje:
  - backend/database/seeders/data/optokom.json   (rekordy do seedera)
  - backend/database/seeders/data/optokom_mail_draft.txt (draft maila — NIE wysyłany)
  - raport na stdout

Uruchom: python3 scripts/import_optokom.py
"""
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = "/home/dev/Pobrane/LISTA LOKALIZACJI 2026 REZERWACJE_1.xlsx"
OUT_DIR = os.path.join(ROOT, "backend/database/seeders/data")
OUT_JSON = os.path.join(OUT_DIR, "optokom.json")
OUT_MAIL = os.path.join(OUT_DIR, "optokom_mail_draft.txt")
GEOCACHE = os.path.join(OUT_DIR, "optokom_geocache.json")

OWNER_EMAIL = "biuro@optokom.pl"
PHONE = "666 926 666"
PLACEHOLDER = "advertisements/optokom-placeholder.jpg"
CURRENT_MONTH_IDX = 5  # czerwiec (sty=0 ... gru=11); dziś 2026-06

# --- Region per miasto (po normalizacji) ---
MALOPOLSKIE = {"Chrzanów", "Trzebinia", "Bukowno"}
MAZOWIECKIE = {"Warszawa", "Ożarów Mazowiecki", "Józefów", "Karczew", "Ostrowik"}
LUBELSKIE = {"Biała Podlaska"}

CITY_NORM = {
    "Dąbrowa Gór.": "Dąbrowa Górnicza",
    "Ożarów Maz.": "Ożarów Mazowiecki",
    "Myslowice": "Mysłowice",
    "Chrzanów – Balin": "Chrzanów",
    "Poręba/Zawiercie": "Poręba",
}

# --- Cennik agencji: tier -> (papier|backlight) -> (najem, montaż, demontaż, druk) netto/mc ---
PRICING = {
    "18":   {"paper": (2300, 200, 200, 270),  "backlight": (3500, 350, 350, 1080)},
    "12":   {"paper": (950, 150, 150, 180),   "backlight": (2000, 350, 350, 720)},
    "elew": {"paper": (5500, 1200, 800, 1800), "backlight": (7500, 3600, 800, 1800)},
}
MONTHS_PL = ["styczeń", "luty", "marzec", "kwiecień", "maj", "czerwiec",
             "lipiec", "sierpień", "wrzesień", "październik", "listopad", "grudzień"]


def norm_city(raw: str) -> str:
    c = (raw or "").strip()
    return CITY_NORM.get(c, c)


def region_for(city: str) -> str:
    if city in MALOPOLSKIE:
        return "małopolskie"
    if city in MAZOWIECKIE:
        return "mazowieckie"
    if city in LUBELSKIE:
        return "lubelskie"
    return "śląskie"


def parse_format(fmt: str):
    """'6,00 x 3,00 (dwustronna)' -> (6.0, 3.0, dwustronna_bool)"""
    fmt = (fmt or "").strip()
    double = "dwustronna" in fmt.lower()
    core = fmt.lower().replace("(dwustronna)", "").strip()
    parts = core.replace(",", ".").split("x")
    try:
        w = float(parts[0].strip())
        h = float(parts[1].strip())
    except (ValueError, IndexError):
        return None, None, double
    return w, h, double


def price_tier(area: float):
    if 15 <= area <= 21:
        return "18"
    if 9 <= area <= 14:
        return "12"
    if 55 <= area <= 65:
        return "elew"
    return None  # nietypowy → wycena indywidualna


def road_class_for(text: str) -> str:
    t = text.upper()
    for marker in ("DK86", "DK 86", "DK94", "DK 94", "DK 94", "E86", "TRASA E", " DK ", "DK94"):
        if marker in t:
            return "national"
    return "urban"


GEO = {}
def load_geocache():
    global GEO
    if os.path.exists(GEOCACHE):
        with open(GEOCACHE, encoding="utf-8") as f:
            GEO = json.load(f)


def save_geocache():
    with open(GEOCACHE, "w", encoding="utf-8") as f:
        json.dump(GEO, f, ensure_ascii=False, indent=2)


def clean_street(s: str) -> str:
    """'ul. Tetmajera 86/ul. Krakowska (Rondo)' -> 'Tetmajera 86'"""
    s = re.sub(r"(?i)\bul\.?\s*", "", s)       # usuń 'ul.'
    s = re.sub(r"\(.*?\)", "", s)              # usuń nawiasy
    s = s.split("/")[0]                         # pierwsza ulica przed '/'
    return s.strip(" ,")


def _nominatim(query: str):
    q = urllib.parse.quote(query)
    url = f"https://nominatim.openstreetmap.org/search?format=json&limit=1&q={q}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "reklamap-import/1.0 (kontakt@reklamap.pl)"})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.load(r)
        time.sleep(1.1)  # rate limit Nominatim
        if data:
            return [round(float(data[0]["lat"]), 6), round(float(data[0]["lon"]), 6)]
    except Exception as e:
        print(f"   ⚠ Nominatim '{query}': {e}", file=sys.stderr)
    return None


def geocode_city(city: str):
    """Centroid miasta — fallback, gdy ulicy nie da się zlokalizować."""
    key = f"__city__|{city}"
    if key in GEO and GEO[key] is not None:
        return GEO[key]
    res = _nominatim(f"{city}, Polska")
    GEO[key] = res
    return res


def geocode_location(loc: str, city_fallback: str):
    """Geokoduje z opisu lokalizacji — wyłuskuje właściwą miejscowość, gdy
    kolumna 'Miasto' jest zbiorcza (np. cały powiat jako 'Biała Podlaska',
    a nośnik stoi w Konstantynowie/Terespolu/Janowie Podlaskim)."""
    key = f"loc|{loc}"
    if key in GEO and GEO[key] is not None:
        return GEO[key]
    # miejscowość = wiodące słowa przed ulicą/skrzyżowaniem/nawiasem/numerem
    # ("Wisznice (Rond) sk..." → "Wisznice"; "Konstantynów, ul..." → "Konstantynów")
    m_loc = re.match(r"\s*([A-Za-zĄĆĘŁŃÓŚŹŻąćęłńóśźż][A-Za-zĄĆĘŁŃÓŚŹŻąćęłńóśźż\s-]*?)(?=\s*(\(|,|/|\bul\.?|\bsk\.?|\d|$))", loc)
    locality = m_loc.group(1).strip() if m_loc and m_loc.group(1).strip() else None
    m = re.search(r"(?i)ul\.?\s*([^,/()]+)", loc)
    street_name = m.group(1).strip() if m else None
    cands = []
    if street_name and locality:
        cands.append(f"{street_name}, {locality}, Polska")
    if locality:
        cands.append(f"{locality}, Polska")
    if street_name:
        cands.append(f"{street_name}, {city_fallback}, Polska")
    res = None
    for q in cands:
        res = _nominatim(q)
        if res:
            break
    GEO[key] = res
    return res


def geocode(street: str, city: str):
    key = f"{street}|{city}"
    if key in GEO and GEO[key] is not None:
        return GEO[key]
    clean = clean_street(street)
    no_num = re.sub(r"\s*\d.*$", "", clean).strip()  # ulica bez numeru
    # Strategia: oczyszczona ulica → ulica bez numeru. (Bez centroidu miasta —
    # niedokładny pin myli; nieznalezione zostają draftem do ręcznego uzupełnienia.)
    res = None
    for q in (f"{clean}, {city}, Polska", f"{no_num}, {city}, Polska"):
        if q.strip().startswith(","):
            continue
        res = _nominatim(q)
        if res:
            break
    GEO[key] = res
    return res


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    load_geocache()
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["LOKALIZACJE"]

    records = []
    stats = {"active": 0, "soon_available": 0, "reserved": 0, "draft": 0,
             "geocoded": 0, "no_coords": 0, "price_unknown": 0, "approx": 0}
    drafts = []

    for r in range(3, ws.max_row + 1):
        num = ws.cell(r, 1).value
        fmt_raw = ws.cell(r, 6).value
        if not num or not fmt_raw:
            continue
        num = str(num).strip()
        city = norm_city(ws.cell(r, 2).value)
        street = (ws.cell(r, 3).value or "").strip()
        loc_desc = (ws.cell(r, 4).value or "").strip()
        direction = (ws.cell(r, 5).value or "").strip()
        oswietlenie = (ws.cell(r, 7).value or "").strip()
        material = (ws.cell(r, 8).value or "").strip()
        gps_raw = (str(ws.cell(r, 9).value or "")).strip()
        resv = [ws.cell(r, 10 + i).value for i in range(12)]

        w, h, double = parse_format(str(fmt_raw))
        if not w or not h:
            continue
        area = w * h

        is_backlight_tech = material.upper() == "BACKLIGHT"
        has_backlight = is_backlight_tech or oswietlenie.upper() in ("TAK", "BACKLIGHT")
        ad_type = "wall" if (max(w, h) > 6.5 or area > 30) else "billboard"
        orientation = "landscape" if w >= h else "portrait"

        # --- Cena z cennika ---
        tier = price_tier(area)
        if tier:
            kind = "backlight" if is_backlight_tech else "paper"
            rent, mount, demount, printc = PRICING[tier][kind]
        else:
            rent = mount = demount = printc = 0
            stats["price_unknown"] += 1

        # --- Współrzędne ---
        lat = lon = None
        if gps_raw:
            try:
                parts = gps_raw.split(",")
                lat = round(float(parts[0].strip()), 6)
                lon = round(float(parts[1].strip()), 6)
            except (ValueError, IndexError):
                lat = lon = None
        approx = False
        if lat is None and street:
            geo = geocode(street, city)
            if geo:
                lat, lon = geo
                stats["geocoded"] += 1
        if lat is None:
            # właściwa miejscowość bywa w opisie LUB w polu ulicy (kolumna 'Miasto'
            # potrafi być zbiorcza — cały powiat). Próbujemy obu źródeł.
            for src in (street, loc_desc):
                if not src:
                    continue
                geo = geocode_location(src, city)
                if geo:
                    lat, lon = geo
                    stats["geocoded"] += 1
                    break
        if lat is None:
            # nie znaleziono dokładnego punktu → pinezka orientacyjna w podanej
            # okolicy (centroid miasta). Nośnik ZOSTAJE widoczny, z adnotacją.
            c = geocode_city(city)
            if c:
                lat, lon = c
                approx = True
                stats["approx"] += 1

        # --- Status z kalendarza ---
        available_from = None
        if tier is None:
            status, is_active = "draft", False
            stats["draft"] += 1
            drafts.append(f"{num} ({city}, {w:g}×{h:g} m, {area:.0f} m², {material})")
        elif lat is None:
            # nawet centroidu nie udało się ustalić → draft (nie powinno się zdarzyć)
            status, is_active = "draft", False
            stats["draft"] += 1
            stats["no_coords"] += 1
        elif not resv[CURRENT_MONTH_IDX]:
            status, is_active = "active", True
            stats["active"] += 1
        else:
            nxt = next((i for i in range(CURRENT_MONTH_IDX + 1, 12) if not resv[i]), None)
            if nxt is not None:
                status, is_active = "soon_available", True
                available_from = f"2026-{nxt + 1:02d}-01"
                stats["soon_available"] += 1
            else:
                status, is_active = "reserved", True
                stats["reserved"] += 1

        # --- Tytuł + opis ---
        type_label = "Billboard" if ad_type == "billboard" else "Powierzchnia reklamowa na elewacji"
        loc_bits = " – ".join([b for b in [street, loc_desc] if b])
        title = f"{type_label} {w:g}×{h:g} m – {city}"
        if loc_desc:
            title += f", {loc_desc}"
        title = title[:120]

        desc_lines = []
        desc_lines.append(
            f"Nośnik reklamowy ({material or 'standard'}) {w:g}×{h:g} m, powierzchnia {area:.0f} m², "
            f"w lokalizacji: {loc_bits or city} ({city})."
        )
        if direction:
            desc_lines.append(f"Widoczność w kierunku: {direction}.")
        if has_backlight:
            desc_lines.append("Nośnik podświetlany.")
        if double:
            desc_lines.append("Nośnik dwustronny.")
        if tier:
            druk_txt = f"{printc} zł" if printc else "wycena"
            desc_lines.append(
                f"Cennik (ceny netto): najem {rent} zł/mc · montaż {mount} zł · "
                f"demontaż {demount} zł · druk {druk_txt} (montaż, demontaż i druk jednorazowo)."
            )
        else:
            desc_lines.append("Powierzchnia nietypowa — wycena indywidualna.")
        # Nr nośnika (np. BE04) to wewnętrzny ref agencji — trzymany w polu
        # "ref" w optokom.json do powiązania z arkuszem, NIE w publicznym opisie.
        if approx:
            desc_lines.append("Lokalizacja orientacyjna — dokładny punkt do potwierdzenia.")
        desc_lines.append("Zdjęcie poglądowe — rzeczywista fotografia nośnika zostanie dodana wkrótce.")
        description = "\n".join(desc_lines)

        rec = {
            "ref": num,
            "title": title,
            "type": ad_type,
            "location": (street or loc_desc or city)[:255],
            "city": city,
            "region": region_for(city),
            "latitude": lat,
            "longitude": lon,
            "description": description,
            "price": rent,
            "price_unit": "month",
            "width": w,
            "height": h,
            "orientation": orientation,
            "has_backlight": has_backlight,
            "has_image": True,
            "image_url": PLACEHOLDER,
            "images": [PLACEHOLDER],
            "owner_email": OWNER_EMAIL,
            "phone": PHONE,
            "offer_type": "agency",
            "has_vat_invoice": True,
            "price_includes_print": False,
            "price_includes_mounting": False,
            "price_negotiable": False,
            "graphic_design_help": False,
            "traffic_type": ["vehicular"],
            "status": status,
            "is_active": is_active,
            "available_from": available_from,
        }
        if ad_type == "billboard":
            rec["variant"] = "standard"
            rec["road_class"] = road_class_for(f"{street} {loc_desc} {direction}")
        records.append(rec)

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    save_geocache()

    # --- Draft maila do agencji (NIE wysyłany) ---
    live = stats["active"] + stats["soon_available"]
    mail = f"""DRAFT — do akceptacji przed wysłaniem (NIE wysłany automatycznie)
Do: {OWNER_EMAIL}
Temat: Państwa nośniki na ReklaMap — podsumowanie i kilka pytań

Dzień dobry,

dziękujemy za przesłaną listę lokalizacji. Dodaliśmy Państwa nośniki do
serwisu reklamap.pl — łącznie {len(records)}, z czego {live} jest już widocznych
(dostępne i wkrótce dostępne), reszta to nośniki zarezerwowane oraz pozycje
oczekujące na wycenę.

Żeby ceny w serwisie były kompletne, prosimy o doprecyzowanie:

1. Druk dla nośników typu BANNER VINYL i Siatka Mesh — jaki koszt? (cennik
   podawał druk tylko dla nośników klejonych papierem i Backlight)
2. Powierzchnie nietypowe — "wycena indywidualna" (czekają u nas jako szkice):
{chr(10).join('   - ' + d for d in drafts) if drafts else '   (brak)'}

Pozostałe nośniki mają już ceny zgodne z przesłanym cennikiem.

W razie zmian w portfolio lub aktualizacji dostępności wystarczy odpisać
na tego maila — zajmiemy się aktualizacją.

Pozdrawiamy,
zespół ReklaMap
"""
    with open(OUT_MAIL, "w", encoding="utf-8") as f:
        f.write(mail)

    # --- Raport ---
    print(f"\n{'='*60}\nIMPORT OPTOKOM — RAPORT\n{'='*60}")
    print(f"Rekordów:            {len(records)}")
    print(f"  active (dostępne): {stats['active']}")
    print(f"  soon_available:    {stats['soon_available']}")
    print(f"  reserved:          {stats['reserved']}")
    print(f"  draft:             {stats['draft']}  (w tym bez współrzędnych: {stats['no_coords']}, wycena indywidualna: {stats['price_unknown']})")
    print(f"Geokodowano braków:  {stats['geocoded']}")
    print(f"Pinezka orientacyjna (widoczne, z adnotacją): {stats['approx']}")
    print(f"\nWidoczne publicznie (active+soon): {live}")
    print(f"\nJSON:        {OUT_JSON}")
    print(f"Draft maila: {OUT_MAIL}  (NIE wysłany)")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
