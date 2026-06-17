#!/usr/bin/env python3
"""Import nośników agencji Big Group Sp. z o.o. Sk. z cenników Excel + zdjęć ze stron produktowych.

Źródło: /home/dev/Pobrane/biggroup/*.xlsx (6 cenników: Autostrada A2, Wrocław,
Poznań, Września, Gniezno, Kościan).

Mapowanie typów (ustalone po oględzinach zdjęć ze stron produktowych biggroup.pl):
  - "Wielki Format" miejski (siatki wielkoformatowe na elewacjach)      → wall
  - Autostrada A2 (wolnostojące tablice przy autostradzie)              → billboard / highway
  - Backlight (podświetlane gabloty w przejściach dworcowych)           → citylight / indoor
  - Citylight 1,2×1,8 m                                                 → citylight / indoor

Cena główna = stawka najmu za 1 miesiąc (z rabatem, gdzie cennik go podaje),
jednostka /miesiąc. Pełna tabela progów (1/3/6+ mc) + druk + montaż trafia do opisu.
Druk zawsze osobno (price_includes_print=False). Montaż wliczony w najem dla nośników
miejskich (footnote „w cenie najmu montaż i demontaż"); przy A2 montaż jest osobną
pozycją cennika → price_includes_mounting=False.

Zdjęcia: pobiera og:image ze strony produktowej każdego nośnika, skaluje i zapisuje
jpg+webp do storage. Zapisuje backend/database/seeders/data/biggroup.json.

Uruchom:  python3 scripts/import_biggroup.py
Potem:    php artisan db:seed --class=BigGroupSeeder
"""

import io
import json
import os
import re
import sys
import time
import urllib.request
from urllib.parse import urlsplit, urlunsplit, quote

import openpyxl
from PIL import Image

SRC_DIR = "/home/dev/Pobrane/biggroup"
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_JSON = os.path.join(ROOT, "backend/database/seeders/data/biggroup.json")
STORE_DIR = os.path.join(ROOT, "backend/storage/app/public/advertisements/biggroup")
REL_PREFIX = "advertisements/biggroup"

OWNER_EMAIL = "info@biggroup.pl"
# Telefonu NIE wstawiamy — agencja nie poprosiła o jego publikację (prosili tylko
# o profil na info@biggroup.pl). Kontakt idzie przez formularz na adres profilu.
# Numery z cenników (602702648 Filip, 784784244 A2) — dopiero po potwierdzeniu mailem.

UA = "Mozilla/5.0 (compatible; reklamap-import/1.0; +kontakt@reklamap.pl)"
FETCH_IMAGES = "--no-images" not in sys.argv

MONTHS_ROMAN = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6,
                "VII": 7, "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12}


# ---------- parsowanie pomocnicze ----------

def _num(x):
    """1234.50 -> '1234,5'; 12.0 -> '12' (przecinek dziesiętny, bez zer końcowych)."""
    if x is None:
        return ""
    s = f"{x:.2f}".rstrip("0").rstrip(".")
    return s.replace(".", ",")


def to_float(v):
    if v is None:
        return None
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_dims(fmt: str):
    """'12,15 x 11,5' / '5 x 12' -> (12.15, 11.5). Pomija formaty złożone (cały nośnik)."""
    if not fmt:
        return None, None
    parts = re.split(r'[x×]', fmt, flags=re.IGNORECASE)
    if len(parts) != 2:
        return None, None
    w = to_float(parts[0])
    h = to_float(parts[1])
    return w, h


def parse_gps(s: str):
    if not s:
        return None, None
    m = re.search(r'(\d{2}\.\d{4,})\s*[, ]\s*(\d{2}\.\d{4,})', str(s))
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def parse_available_from(text: str):
    """Zwraca 'YYYY-MM-DD' jeśli dostępność wskazuje przyszłą datę startu, inaczej None."""
    if not text:
        return None
    t = str(text)
    # 'od 1.09.2026' / 'od 01.02.2027'
    m = re.search(r'od\s+(\d{1,2})\.(\d{1,2})\.(\d{4})', t)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}"
    # 'od X 2026' (miesiąc rzymski)
    m = re.search(r'od\s+([IVX]+)\s+(\d{4})', t)
    if m and m.group(1) in MONTHS_ROMAN:
        return f"{int(m.group(2)):04d}-{MONTHS_ROMAN[m.group(1)]:02d}-01"
    return None


def is_sold_out(text: str) -> bool:
    t = (text or "").lower()
    return "sprzedan" in t and "dostępn" not in t


# ---------- typ nośnika z numeru katalogowego ----------

def classify(catalog: str):
    """catalog -> (type, environment, is_highway, is_backlight)."""
    c = (catalog or "").upper()
    if "A2" in c:
        return "billboard", "outdoor", True, False
    if c.startswith("BL-"):
        return "citylight", "indoor", False, True
    if c.startswith("CL-"):
        return "citylight", "indoor", False, True
    # WF-<miasto> bez A2 = siatka wielkoformatowa na elewacji
    return "wall", "outdoor", False, False


# ---------- zdjęcia ----------

def safe_url(u: str) -> str:
    """Percent-encode znaki nie-ASCII w URL (og:image bywa z polskimi znakami)."""
    p = urlsplit(u)
    return urlunsplit((p.scheme, p.netloc, quote(p.path), quote(p.query, safe="=&?"), p.fragment))


def fetch_og_image(product_url: str):
    if not product_url or "/product/" not in product_url:
        return None
    try:
        req = urllib.request.Request(safe_url(product_url), headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=30) as r:
            html = r.read().decode("utf-8", "ignore")
        m = re.search(r'<meta\s+property=["\']og:image["\']\s+content=["\']([^"\']+)["\']', html, re.I)
        return m.group(1) if m else None
    except Exception as e:
        print(f"    og:image błąd: {e}")
        return None


def save_photo(img_url: str, ref: str):
    try:
        req = urllib.request.Request(safe_url(img_url), headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=40) as r:
            data = r.read()
        img = Image.open(io.BytesIO(data)).convert("RGB")
        if img.width > 1600:
            img = img.resize((1600, round(img.height * 1600 / img.width)))
        os.makedirs(STORE_DIR, exist_ok=True)
        img.save(os.path.join(STORE_DIR, f"{ref}.jpg"), "JPEG", quality=85)
        img.save(os.path.join(STORE_DIR, f"{ref}.webp"), "WEBP", quality=85)
        return f"{REL_PREFIX}/{ref}.jpg"
    except Exception as e:
        print(f"    foto błąd {ref}: {e}")
        return None


# ---------- odczyt arkuszy z hiperłączami ----------

def load_sheet(path, sheet=None):
    wb_v = openpyxl.load_workbook(path, data_only=True)
    wb_l = openpyxl.load_workbook(path)
    ws_v = wb_v[sheet] if sheet else wb_v.active
    ws_l = wb_l[sheet] if sheet else wb_l.active
    rows = [[c for c in r] for r in ws_v.iter_rows(values_only=True)]
    # mapa hiperłączy: row(1-based) -> product url (pierwszy /product/ w wierszu)
    prod = {}
    for row in ws_l.iter_rows():
        for c in row:
            if c.hyperlink and c.hyperlink.target and "/product/" in c.hyperlink.target:
                prod.setdefault(c.row, c.hyperlink.target)
    return rows, prod


def row_has_product(prod, rownum):
    return prod.get(rownum)


def strcells(row):
    return [str(c).strip() for c in row if isinstance(c, str) and str(c).strip()]


def atoms(cells):
    """Rozbij komórki lokalizacji na atomy (w cennikach lokalizacja bywa w jednej
    komórce rozdzielona ' | ') i odfiltruj szum (LINK, #VALUE!, dostępność, NOŚNIK)."""
    out = []
    for c in cells:
        for part in re.split(r'\s*\|\s*', c):
            p = part.strip()
            if not p:
                continue
            up = p.upper()
            low = p.lower()
            if up in ("LINK", "#VALUE!") or "NOŚNIK" in up:
                continue
            if any(k in low for k in ("dostępn", "sprzedan", "rezerwacja", "osób")):
                continue
            if "zdjęcia" in low:
                continue
            out.append(p)
    return out


def numcells(row):
    out = []
    for c in row:
        if isinstance(c, (int, float)):
            out.append(float(c))
    return out


# ---------- parser TALL (Autostrada, Wrocław, Gniezno, Września) ----------

def parse_tall(path, cfg):
    rows, prod = load_sheet(path, cfg.get("sheet"))
    surfaces = []
    cur = None
    for i, row in enumerate(rows, start=1):
        purl = row_has_product(prod, i)
        if purl:
            # nowy nośnik — wiersz nagłówkowy
            if cur:
                surfaces.append(cur)
            ss = strcells(row)
            avail = next((s for s in ss if "dostępn" in s.lower() or "sprzedan" in s.lower()), "")
            loc_parts = atoms(ss)
            cur = {
                "product_url": purl,
                "loc_parts": loc_parts,
                "prices": numcells(row),
                "availability": avail,
                "sold_out": is_sold_out(avail),
                "traffic": None,
                "catalog": None, "fmt": None, "area": None, "gps": None,
            }
            continue
        if not cur:
            continue
        # wiersze podrzędne: katalog / FORMAT / POWIERZCHNIA / GPS
        new_block = False
        for idx, c in enumerate(row):
            if not isinstance(c, str):
                continue
            cs = c.strip()
            if re.match(r'^(WF|BL|CL)-[A-Z0-9\-]+$', cs, re.I):
                if cur["catalog"] is None:
                    cur["catalog"] = cs
                else:
                    # drugi numer katalogowy w obrębie nośnika = kolejny blok bez
                    # linku produktowego (np. „cały nośnik" ABC) — zamknij bieżący,
                    # pomiń aż do następnego nagłówka z linkiem /product/.
                    new_block = True
                    break
            elif "FORMAT" in cs.upper():
                cur["fmt"] = _next_val(row, idx)
            elif "POWIERZCHNIA" in cs.upper():
                cur["area"] = to_float(_next_val(row, idx))
            elif "GPS" in cs.upper():
                cur["gps"] = _next_val(row, idx)
        if new_block:
            surfaces.append(cur)
            cur = None
    if cur:
        surfaces.append(cur)
    return [_finalize(s, cfg) for s in surfaces]


def _next_val(row, idx):
    for c in row[idx + 1:]:
        if c is not None and str(c).strip():
            return str(c).strip()
    return None


# ---------- parser WIDE (Poznań, Kościan) ----------

def parse_wide(path, cfg):
    rows, prod = load_sheet(path, cfg.get("sheet"))
    surfaces = []
    for i, row in enumerate(rows, start=1):
        purl = row_has_product(prod, i)
        if not purl:
            continue
        ss = strcells(row)
        catalog = next((s for s in ss if re.match(r'^(WF|BL|CL)-[A-Z0-9\-]+$', s, re.I)), None)
        fmt = next((s for s in ss if re.search(r'\d\s*[x×]\s*\d', s)), None)
        gps = next((s for s in ss if re.search(r'\d{2}\.\d{4,}', s)), None)
        traffic = next((s for s in ss if "osób" in s.lower()), None)
        avail_cells = [s for s in ss if any(k in s.lower()
                       for k in ("dostępn", "sprzedan", "rezerwacja"))]
        sold_out = bool(avail_cells) and all("sprzedan" in a.lower() for a in avail_cells)
        partial = any("sprzedan" in a.lower() or "rezerwacja" in a.lower() for a in avail_cells)
        loc_cells = [s for s in ss
                     if s != catalog and s != fmt and s != gps and s != traffic
                     and s not in avail_cells]
        loc_parts = atoms(loc_cells)
        nums = numcells(row)
        # WIDE: numerics = [area, ...prices]
        area = nums[0] if nums else None
        prices = nums[1:]
        surfaces.append({
            "product_url": purl, "loc_parts": loc_parts, "prices": prices,
            "availability": ("Część terminów zarezerwowana — dostępność potwierdza operator"
                             if partial and not sold_out else ""),
            "sold_out": sold_out, "traffic": traffic, "catalog": catalog, "fmt": fmt,
            "area": area, "gps": gps,
        })
    return [_finalize(s, cfg) for s in surfaces]


# ---------- finalizacja rekordu ----------

def _clean_city(token: str) -> str:
    return re.sub(r'\s*k\.\s*.*$', '', token).strip()


def pick_locality(parts):
    """Wybierz miejscowość z części lokalizacji A2 (pomija odcinek/kierunek/stronę)."""
    for p in parts:
        u = p.strip()
        if not u or u.startswith("("):
            continue
        low = u.lower()
        if low.startswith("kier") or low in ("front", "tył", "tyl", "bok"):
            continue
        if " - " in u and u.upper() == u:   # odcinek, np. „POZNAŃ - KONIN"
            continue
        return _clean_city(u)
    return _clean_city(parts[0]) if parts else ""


def _finalize(s, cfg):
    catalog = s["catalog"] or ""
    # pomiń „cały nośnik" (ABC) — brak strony produktowej, redundancja
    if catalog.upper().endswith("ABC"):
        return None
    if any("cały nośnik" in p.lower() for p in s["loc_parts"]):
        return None

    typ, env, is_highway, is_backlight = classify(catalog)
    w, h = parse_dims(s["fmt"])
    lat, lng = parse_gps(s["gps"])
    area = s["area"]

    # Źródłowy błąd cennika: komórka FORMAT niektórych boków A2 zawiera format
    # „całego nośnika". Wszystkie boki A2 są jednolite: 2,4 × 11,5 m (27,6 m²).
    if (w is None or h is None) and catalog.upper().endswith("C") and is_highway:
        w, h, area = 2.4, 11.5, 27.6

    dims_str = f"{_num(w)}×{_num(h)} m" if (w and h) else ""

    loc_parts = s["loc_parts"]
    if cfg.get("city"):
        city = cfg["city"]
        location = ", ".join(loc_parts) if loc_parts else city
    else:
        # Autostrada: [Odcinek, miejscowość, kierunek..., strona]
        city = pick_locality(loc_parts)
        location = ", ".join(loc_parts)

    # ceny wg konfiguracji kolumn
    price_map = dict(zip(cfg["price_cols"], s["prices"])) if s["prices"] else {}
    head = price_map.get("c1") or price_map.get("katalog")

    # kierunek ruchu
    joined = " ".join(loc_parts).lower()
    tdir = []
    if "wjazd" in joined and "wyjazd" in joined:
        tdir = ["both"]
    elif "wjazd" in joined:
        tdir = ["entry"]
    elif "wyjazd" in joined:
        tdir = ["exit"]

    rec = {
        "ref": catalog,
        "typ_meta": {"type": typ, "env": env, "highway": is_highway, "backlight": is_backlight},
        "title": _make_title(typ, city, loc_parts, dims_str),
        "type": typ,
        "location": location,
        "city": city,
        "latitude": lat,
        "longitude": lng,
        "price": head,
        "price_unit": "month",
        "width": w,
        "height": h,
        "area": area,
        "orientation": ("landscape" if (w or 0) >= (h or 0) else "portrait") if (w and h) else "landscape",
        "traffic_intensity": "high",
        "traffic_direction": tdir,
        "traffic_type": (["vehicular"] if typ in ("billboard", "wall") else ["pedestrian"]),
        "has_backlight": is_backlight,
        "price_includes_print": False,
        "price_includes_mounting": not is_highway,  # A2: montaż osobno
        "graphic_design_help": False,
        "price_negotiable": False,
        "has_vat_invoice": True,
        "owner_email": OWNER_EMAIL,
        "phone": None,
        "contact_preference": "form",
        "offer_type": "agency",
        "status": "reserved" if s.get("sold_out") else "active",
        "is_active": True,
        "available_from": parse_available_from(s["availability"]),
        "_traffic": s.get("traffic"),
        "_area": area,
        "_product_url": s["product_url"],
        # pola pomocnicze do opisu (usuwane przed zapisem JSON? — zostają, seeder zignoruje przez fillable? NIE)
        "_price_map": price_map,
        "_availability": s["availability"],
        "_price_labels": cfg["price_labels"],
        "_is_highway": is_highway,
    }
    # typ-zależne pola
    if typ == "billboard":
        rec["variant"] = "standard"
        rec["road_class"] = "highway" if is_highway else "urban"
    elif typ == "citylight":
        rec["variant"] = "single"
        rec["environment"] = "indoor"
    elif typ == "wall":
        rec["lighting_type_banner"] = "none"
        rec["environment"] = "outdoor"

    rec["description"] = _make_description(rec)
    return rec


def _make_title(typ, city, loc_parts, dims_str):
    dims = (dims_str[:-2].strip() if dims_str else "")  # bez końcówki ' m'
    # pierwszy atom niebędący kierunkiem/stroną — opis miejsca
    place = next((p for p in loc_parts
                  if not p.lower().startswith("kier")
                  and p.lower() not in ("front", "tył", "tyl", "bok")
                  and not p.startswith("(")), city)
    dpart = f" {dims} m" if dims else ""
    if typ == "billboard":
        kier = next((p for p in loc_parts if p.lower().startswith("kier")), "")
        side = next((p for p in loc_parts if p.lower() in ("front", "tył", "tyl", "bok")), "")
        kier = re.sub(r'(?i)^kierunek', 'kier.', kier).strip()
        tail = ", ".join(x for x in (kier, side) if x)
        return f"Billboard przy A2{dpart} – {city}" + (f", {tail}" if tail else "")
    if typ == "citylight":
        return f"Powierzchnia podświetlana{dpart} – {city}, {place}"
    return f"Powierzchnia wielkoformatowa{dpart} – {city}, {place}"


def _make_description(rec):
    pm = rec["_price_map"]
    labels = rec["_price_labels"]
    lines = []
    typ = rec["type"]

    if typ == "billboard":
        intro = (f"Wolnostojąca tablica wielkoformatowa przy autostradzie A2 "
                 f"({rec['location']}). Nośnik w ciągu ruchu o bardzo dużym natężeniu, "
                 f"doskonale widoczny dla kierowców na długim, prostym odcinku trasy.")
    elif typ == "citylight":
        intro = (f"Podświetlana powierzchnia reklamowa ({rec['location']}). "
                 f"Lokalizacja o wysokim natężeniu ruchu pieszego, ekspozycja przez całą dobę "
                 f"dzięki podświetleniu.")
    else:
        intro = (f"Powierzchnia wielkoformatowa (siatka/baner) na elewacji budynku "
                 f"({rec['location']}). Duża, dobrze eksponowana płaszczyzna widoczna z dużej "
                 f"odległości — idealna do kampanii wizerunkowych.")
    lines.append(intro)

    if rec.get("width") and rec.get("height"):
        a = f" (ok. {rec['_area']:.0f} m²)" if rec.get("_area") else ""
        lines.append(f"Format: {_num(rec['width'])} × {_num(rec['height'])} m{a}.")

    # tabela cen najmu
    tier_lines = []
    for key, lab in labels:
        if key in pm and pm[key]:
            tier_lines.append(f"  • {lab}: {pm[key]:,.0f} zł netto / mc".replace(",", " "))
    if tier_lines:
        lines.append("Najem (im dłuższa umowa, tym niższa stawka miesięczna):")
        lines.extend(tier_lines)

    extras = []
    if pm.get("druk"):
        extras.append(f"druk grafiki: {pm['druk']:,.0f} zł netto".replace(",", " "))
    if rec["_is_highway"]:
        if pm.get("montaz"):
            extras.append(f"montaż i demontaż: {pm['montaz']:,.0f} zł netto".replace(",", " "))
    else:
        extras.append("montaż, demontaż i utylizacja w cenie najmu")
    if extras:
        lines.append("Koszty dodatkowe: " + "; ".join(extras) + ".")

    lines.append("Wszystkie ceny netto — należy doliczyć VAT 23%.")
    if rec.get("_traffic"):
        lines.append(f"Szacowany miesięczny ruch w lokalizacji: {rec['_traffic']}.")
    if rec["_availability"]:
        lines.append(f"Dostępność: {rec['_availability']}.")
    return "\n".join(lines)


# ---------- konfiguracja plików ----------

def find_file(keyword):
    for f in os.listdir(SRC_DIR):
        if f.endswith(".xlsx") and keyword in f and not f.startswith("."):
            return os.path.join(SRC_DIR, f)
    raise FileNotFoundError(keyword)


FILES = [
    {"key": "Autostrada", "layout": "tall", "city": None,
     "price_cols": ["katalog", "c3", "c6", "montaz", "druk"],
     "price_labels": [("katalog", "umowa 1 mc"), ("c3", "umowa 3 mce"),
                      ("c6", "umowa 6+ mcy")]},
    {"key": "Wroc", "layout": "tall", "city": "Wrocław",
     "price_cols": ["c1", "cmid", "c6", "druk"],
     "price_labels": [("c1", "umowa 1 mc"), ("cmid", "umowa 2–3 mce"),
                      ("c6", "umowa 6+ mcy")]},
    {"key": "Gniezno", "layout": "tall", "city": "Gniezno",
     "price_cols": ["c1", "c3", "c6", "druk"],
     "price_labels": [("c1", "umowa 1 mc"), ("c3", "umowa 3 mce"),
                      ("c6", "umowa 6+ mcy")]},
    {"key": "Wrze", "layout": "tall", "city": "Września",
     "price_cols": ["c1", "cmid", "c6", "druk"],
     "price_labels": [("c1", "umowa 1 mc"), ("cmid", "umowa 2–3 mce"),
                      ("c6", "umowa 6+ mcy")]},
    {"key": "Pozna", "layout": "wide", "city": "Poznań", "sheet": "Poznań",
     "price_cols": ["katalog", "c1", "c3", "c6", "druk"],
     "price_labels": [("c1", "umowa 1 mc"), ("c3", "umowa 3 mce"),
                      ("c6", "umowa 6+ mcy")]},
    {"key": "Ko", "layout": "wide", "city": "Kościan",
     "price_cols": ["c1", "c3", "c6", "druk"],
     "price_labels": [("c1", "umowa 1 mc"), ("c3", "umowa 3 mce"),
                      ("c6", "umowa 6+ mcy")]},
]


def main():
    all_recs = []
    for cfg in FILES:
        path = find_file(cfg["key"])
        print(f"\n=== {os.path.basename(path)} ({cfg['layout']}) ===")
        recs = parse_tall(path, cfg) if cfg["layout"] == "tall" else parse_wide(path, cfg)
        recs = [r for r in recs if r]
        print(f"  nośników: {len(recs)}")
        all_recs.extend(recs)

    # ref unikalny do nazwy pliku zdjęcia
    for r in all_recs:
        r["_imgref"] = re.sub(r'[^a-z0-9]+', '-', (r["ref"] or "").lower()).strip("-")

    # zdjęcia
    if FETCH_IMAGES:
        print(f"\n--- pobieranie zdjęć ({len(all_recs)}) ---")
        for r in all_recs:
            ref = r["_imgref"]
            jpg = os.path.join(STORE_DIR, f"{ref}.jpg")
            webp = os.path.join(STORE_DIR, f"{ref}.webp")
            if os.path.exists(jpg) and os.path.exists(webp):
                img_rel = f"{REL_PREFIX}/{ref}.jpg"   # już pobrane — nie ściągaj ponownie
                print(f"  {r['ref']}: cache")
            else:
                og = fetch_og_image(r["_product_url"])
                img_rel = save_photo(og, ref) if og else None
                print(f"  {r['ref']}: {'OK' if img_rel else 'BRAK'}")
                time.sleep(0.4)
            r["image_url"] = img_rel
            r["images"] = [img_rel] if img_rel else []
            r["has_image"] = bool(img_rel)
    else:
        for r in all_recs:
            r["image_url"] = None
            r["images"] = []
            r["has_image"] = False

    # czyść pola pomocnicze
    clean = []
    for r in all_recs:
        c = {k: v for k, v in r.items() if not k.startswith("_") and k not in
             ("ref", "product_url", "area", "typ_meta")}
        clean.append(c)

    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    json.dump(clean, open(OUT_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    with_img = sum(1 for r in all_recs if r.get("has_image"))
    no_geo = sum(1 for r in clean if not r["latitude"])
    by_type = {}
    for r in clean:
        by_type[r["type"]] = by_type.get(r["type"], 0) + 1
    print(f"\nGotowe → {OUT_JSON}")
    print(f"  Nośników: {len(clean)} | ze zdjęciem: {with_img} | bez geo: {no_geo}")
    print(f"  Wg typu: {by_type}")
    print(f"\nTeraz: php artisan db:seed --class=BigGroupSeeder")
    return 0


if __name__ == "__main__":
    sys.exit(main())
